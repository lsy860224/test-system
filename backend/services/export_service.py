from io import BytesIO
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from models.standard import StandardItem, StandardCategory
from models.equipment import Equipment
from models.schedule import TestSchedule
from models.ncr import NCRReport
from models.vendor import VendorLab
from models.sop import SOP
from models.project import Project
from models.customer import Customer
from models.item import Item
from models.user import User
from services.project_service import compute_project_progress
from services.schedule_service import compute_status

HDR_FILL = PatternFill("solid", fgColor="2B2F82")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)


def _sanitize_cell(val):
    """자유 텍스트 필드(NCR 노트, 업체 주소 등)가 그대로 xlsx 셀에 들어가므로,
    수신자가 레거시 Excel에서 열 때 수식으로 해석될 수 있는 값(=+-@로 시작)을 텍스트로 무력화한다."""
    if isinstance(val, str) and val[:1] in ("=", "+", "-", "@"):
        return "'" + val
    return val


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list], col_widths: list[int]):
    ws = wb.create_sheet(title)
    for col, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=_sanitize_cell(val))
    return ws


def generate_full_export(db: Session) -> bytes:
    """전체 모듈 데이터를 시트별로 담은 단일 Excel 워크북 생성 (백업/외부 공유용, 실측 DB 값 그대로)."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── 규격 매트릭스 ──────────────────────────────────────
    categories = {c.id: c.name_ko for c in db.query(StandardCategory).all()}
    standards = db.query(StandardItem).filter(StandardItem.is_deleted == False).order_by(StandardItem.standard_no, StandardItem.standard_code).all()
    _write_sheet(
        wb, "규격매트릭스",
        ["규격 No.", "규격명", "Rev.", "항목 No.", "시험 항목명", "분류", "수행방식", "우선순위",
         "DV 목표일", "DV 완료일", "PV 목표일", "PV 완료일", "메모"],
        [[
            s.standard_no, s.standard_name, s.revision_no, s.standard_code, s.name,
            categories.get(s.category_id, ""), s.source_type, s.priority,
            s.dv_target_date, s.dv_actual_date, s.pv_target_date, s.pv_actual_date, s.notes,
        ] for s in standards],
        [14, 32, 10, 12, 30, 12, 10, 10, 12, 12, 12, 12, 26],
    )

    # ── 장비대장 ────────────────────────────────────────────
    equipment = db.query(Equipment).order_by(Equipment.name).all()
    _write_sheet(
        wb, "장비대장",
        ["장비명", "모델", "제조사", "자산번호", "분류", "담당자", "상태", "위치", "구매일", "메모"],
        [[e.name, e.model, e.manufacturer, e.asset_number, e.category, e.manager, e.status, e.location, e.purchase_date, e.notes] for e in equipment],
        [22, 16, 16, 14, 14, 10, 10, 14, 12, 26],
    )

    # ── 시험 일정 ───────────────────────────────────────────
    proj_names = {p.id: p.name for p in db.query(Project).all()}
    std_names = {s.id: s.name for s in db.query(StandardItem).all()}
    schedules = db.query(TestSchedule).order_by(TestSchedule.planned_start).all()
    _write_sheet(
        wb, "시험일정",
        ["프로젝트", "시험 항목", "시험 유형", "계획 시작", "계획 종료", "실제 시작", "실제 종료", "상태", "결과", "메모"],
        [[
            proj_names.get(sc.project_id, ""), std_names.get(sc.standard_item_id, ""), sc.test_type,
            sc.planned_start, sc.planned_end, sc.actual_start, sc.actual_end, compute_status(sc), sc.result, sc.notes,
        ] for sc in schedules],
        [20, 26, 10, 12, 12, 12, 12, 8, 8, 26],
    )

    # ── NCR ────────────────────────────────────────────────
    ncrs = db.query(NCRReport).order_by(NCRReport.detected_date.desc()).all()
    _write_sheet(
        wb, "NCR",
        ["NCR No.", "부품명", "시험 단락", "이슈 요약", "심각도", "상태", "발견일", "조치기한", "종결일"],
        [[n.ncr_number, n.part_name, n.test_section, n.issue_summary, n.severity, n.status, n.detected_date, n.due_date, n.closed_date] for n in ncrs],
        [14, 20, 14, 36, 8, 10, 12, 12, 12],
    )

    # ── 외주 시험소 ─────────────────────────────────────────
    vendors = db.query(VendorLab).order_by(VendorLab.name).all()
    _write_sheet(
        wb, "외주시험소",
        ["시험소명", "약칭", "유형", "KOLAS", "담당자", "연락처", "이메일", "주소", "운영여부"],
        [[v.name, v.short_name, v.lab_type, "Y" if v.kolas_certified else "N", v.contact_name, v.contact_phone, v.contact_email, v.address, "운영중" if v.is_active else "중지"] for v in vendors],
        [22, 12, 14, 8, 12, 14, 20, 30, 10],
    )

    # ── 절차서 ────────────────────────────────────────────────
    sops = db.query(SOP).order_by(SOP.sop_number).all()
    user_names = {u.id: u.name for u in db.query(User).all()}
    _write_sheet(
        wb, "절차서",
        ["문서번호", "제목", "종류", "버전", "분류", "상태", "작성자", "승인자", "발행일", "최근개정일"],
        [[s.sop_number, s.title, s.doc_type, s.version, s.category, s.status, s.owner,
          user_names.get(s.approver_id), s.issue_date, s.revision_date] for s in sops],
        [16, 32, 12, 10, 14, 10, 12, 12, 12, 12],
    )

    # ── 프로젝트 ────────────────────────────────────────────
    cust_names = {c.id: c.name for c in db.query(Customer).all()}
    item_names = {i.id: i.name for i in db.query(Item).all()}
    projects = db.query(Project).order_by(Project.project_code).all()
    progress_map = compute_project_progress(db, [p.id for p in projects])
    _write_sheet(
        wb, "프로젝트",
        ["프로젝트 코드", "프로젝트명", "고객사", "아이템", "단계", "상태", "시작일", "목표일", "완료일", "진행률"],
        [[
            p.project_code, p.name, cust_names.get(p.customer_id, ""), item_names.get(p.item_id, ""), p.phase, p.status,
            p.start_date, p.target_date, p.actual_date, progress_map.get(p.id, 0),
        ] for p in projects],
        [16, 24, 18, 20, 10, 8, 12, 12, 12, 10],
    )

    # ── 업체 ────────────────────────────────────────────────
    customers = db.query(Customer).order_by(Customer.name).all()
    _write_sheet(
        wb, "업체",
        ["업체명", "약칭", "구분", "사업자번호", "홈페이지", "주소", "운영여부"],
        [[c.name, c.short_name, c.company_type, c.business_reg_number, c.homepage, c.address, "운영중" if c.is_active else "중지"] for c in customers],
        [22, 12, 14, 16, 24, 30, 10],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
