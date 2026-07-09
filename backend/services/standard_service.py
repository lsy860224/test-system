from datetime import datetime, date
from io import BytesIO
from fastapi import HTTPException
from sqlalchemy.orm import Session
from models.standard import StandardItem, StandardCategory, StandardHistory
from schemas.standard import StandardItemCreate, StandardItemUpdate
from services import pagination_helper, sop_coverage

# ── categories ─────────────────────────────────────────────
def list_categories(db: Session) -> list:
    return db.query(StandardCategory).order_by(StandardCategory.display_order).all()

# ── items ───────────────────────────────────────────────────
def list_items(db: Session, page: int, size: int, category_id, source_type, search) -> dict:
    q = db.query(StandardItem).filter(StandardItem.is_deleted == False)
    if category_id:
        q = q.filter(StandardItem.category_id == category_id)
    if source_type:
        q = q.filter(StandardItem.source_type == source_type)
    if search:
        q = q.filter(
            StandardItem.standard_code.ilike(f"%{search}%") |
            StandardItem.name.ilike(f"%{search}%") |
            StandardItem.standard_no.ilike(f"%{search}%") |
            StandardItem.standard_name.ilike(f"%{search}%")
        )
    result = pagination_helper.paginate(q.order_by(StandardItem.standard_code), page, size)
    result["items"] = attach_category_names(db, result["items"])
    attach_sop_coverage(db, result["items"])
    return result

def get_item(db: Session, item_id: int) -> StandardItem:
    item = db.query(StandardItem).filter(StandardItem.id == item_id, StandardItem.is_deleted == False).first()
    if not item:
        raise HTTPException(status_code=404, detail="규격 항목을 찾을 수 없습니다")
    attach_sop_coverage(db, [item])
    return item

def create_item(db: Session, body: StandardItemCreate, created_by: int) -> StandardItem:
    item = StandardItem(**body.model_dump(), created_by=created_by, priority_score=_calc_priority(body.priority))
    db.add(item)
    db.commit()
    db.refresh(item)
    return item

def update_item(db: Session, item_id: int, body: StandardItemUpdate, changed_by: int) -> StandardItem:
    item = get_item(db, item_id)
    changes = _track_changes(item, body.model_dump(exclude_unset=True), changed_by)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    item.updated_at = datetime.utcnow()
    for h in changes:
        db.add(h)
    db.commit()
    db.refresh(item)
    return item

def update_group_info(db: Session, body, changed_by: int) -> dict:
    """규격 그룹(standard_no 동일 항목 전체)의 기본 정보(규격 No./규격명/Revision No.)를 일괄 수정."""
    q = db.query(StandardItem).filter(StandardItem.is_deleted == False)
    q = q.filter(StandardItem.standard_no == body.old_standard_no) if body.old_standard_no else q.filter(StandardItem.standard_no.is_(None))
    items = q.all()
    if not items:
        raise HTTPException(status_code=404, detail="해당 규격 그룹을 찾을 수 없습니다")

    updates = {
        "standard_no": body.standard_no or None,
        "standard_name": body.standard_name,
        "revision_no": body.revision_no,
    }
    for item in items:
        changes = _track_changes(item, updates, changed_by)
        for field, value in updates.items():
            setattr(item, field, value)
        item.updated_at = datetime.utcnow()
        for h in changes:
            db.add(h)
    db.commit()
    return {"updated": len(items)}

def soft_delete(db: Session, item_id: int):
    item = get_item(db, item_id)
    item.is_deleted = True
    db.commit()

def get_history(db: Session, item_id: int) -> list:
    get_item(db, item_id)
    return db.query(StandardHistory).filter(StandardHistory.standard_item_id == item_id).order_by(StandardHistory.changed_at.desc()).all()

# ── Excel template & import ────────────────────────────────
def generate_template() -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "규격매트릭스"

    # 컬럼 순서: 규격 No. | 규격명 | Revision No. | 항목 No.* | 시험항목명* | 분류코드 | 시험조건요약 | 메모
    headers = ["규격 No. *", "규격명", "Revision No.", "항목 No. *", "시험 항목명 *", "분류 코드", "시험 조건 요약", "메모"]
    col_widths = [16, 36, 12, 12, 32, 12, 26, 26]

    hdr_fill = PatternFill("solid", fgColor="2B2F82")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    for col, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 26

    row_fill = PatternFill("solid", fgColor="F5F7FA")
    samples = [
        ["ISO 16750-2", "도로차량 전기전자장치 환경조건 및 시험", "Ed.4.0", "5.1", "저전압 시험", "ELEC", "Vmin=6V, t=60min", ""],
        ["ISO 16750-2", "도로차량 전기전자장치 환경조건 및 시험", "Ed.4.0", "5.2", "고전압 시험", "ELEC", "Vmax=16V, t=60min", ""],
        ["ISO 16750-4", "도로차량 전기전자장치 환경조건 및 시험", "Ed.3.0", "4.1.1", "온도 사이클 시험", "ENV", "-40°C~+85°C, 1000 cycle", ""],
        ["ES 60100-NE30", "HKMC 전기전자 부품 신뢰성 시험", "Rev.5", "6.3.1", "진동 내구 시험", "MECH", "10~2000Hz, 5G", ""],
    ]
    for r, row_data in enumerate(samples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill

    ws2 = wb.create_sheet("작성 가이드")
    ws2["A1"] = "규격 매트릭스 Excel 양식 작성 가이드"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 85
    for i, text in enumerate([
        "■ 규격 No. *: 표준 번호 (예: ISO 16750-2, ES 60100-NE30, CISPR 25)  ← 필수 입력",
        "■ 규격명: 표준의 전체 명칭 (예: 도로차량 전기전자장치 환경조건 및 시험)",
        "■ Revision No.: 표준 개정 번호 (예: Ed.4.0, Rev.5, 2022-01)",
        "■ 항목 No. *: 규격 내 조항 번호 (예: 5.1, 6.3.1)  ← 필수 입력",
        "■ 시험 항목명 *: 시험의 명칭 (예: 온도 사이클 시험, 저전압 시험)  ← 필수 입력",
        "■ 분류 코드: ENV(환경내구) / ELEC(전기전자) / EMC(EMC·EMI) / MECH(기계물리) / REL(신뢰성) / SAFE(기능안전) / PERF(성능기능) / ETC(기타)",
        "■ 시험 조건 요약: 핵심 시험 조건 (예: -40°C~+85°C, 1000 cycle)",
        "■ 메모: 외주 의뢰 조건, 특이사항 등",
        "",
        "※ 같은 규격의 여러 항목은 '규격 No.' 컬럼을 반복 입력하세요.",
        "※ 1행(헤더)은 수정하지 마세요.",
        "※ 수행방식·우선순위·상태·일정은 프로젝트/시험일정 기능에서 관리합니다.",
    ], 3):
        ws2.cell(row=i, column=1, value=text)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def import_from_excel(db: Session, file_bytes: bytes, created_by: int) -> dict:
    """
    Excel 양식 컬럼 순서:
    A: 규격 No.* | B: 규격명 | C: Revision No. | D: 항목 No.* | E: 시험 항목명* | F: 분류 코드 | G: 시험 조건 요약 | H: 메모
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
    except Exception:
        raise HTTPException(status_code=400, detail="올바른 Excel(.xlsx) 파일이 아니거나 파일이 손상되었습니다")
    ws = wb.active

    cats = {c.code.upper(): c.id for c in db.query(StandardCategory).all()}
    existing = {item.standard_code for item in db.query(StandardItem).filter(StandardItem.is_deleted == False).all()}

    created, skipped, errors = [], [], []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(not c for c in row[:5]):
            continue

        standard_no  = str(row[0]).strip() if len(row) > 0 and row[0] else None
        standard_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
        revision_no  = str(row[2]).strip() if len(row) > 2 and row[2] else None
        standard_code = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        name         = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        if not standard_code or not name:
            errors.append(f"행 {row_idx}: 항목 No. 또는 항목명 누락")
            continue

        if standard_code in existing:
            skipped.append(standard_code)
            continue

        cat_raw   = str(row[5]).strip().upper() if len(row) > 5 and row[5] else ""
        cond_raw  = str(row[6]).strip() if len(row) > 6 and row[6] else None
        notes_raw = str(row[7]).strip() if len(row) > 7 and row[7] else None

        item = StandardItem(
            standard_no=standard_no,
            standard_name=standard_name,
            revision_no=revision_no,
            standard_code=standard_code,
            name=name,
            category_id=cats.get(cat_raw),
            test_condition_summary=cond_raw,
            notes=notes_raw,
            created_by=created_by,
        )
        db.add(item)
        created.append(standard_code)
        existing.add(standard_code)

    db.commit()
    return {"created": len(created), "skipped": len(skipped), "errors": errors}


# ── helpers ────────────────────────────────────────────────
def _calc_priority(priority: str) -> int:
    return {"High": 3, "Med": 2, "Low": 1}.get(priority, 2)

def attach_category_names(db: Session, items: list) -> list:
    cats = {c.id: c for c in db.query(StandardCategory).all()}
    for item in items:
        cat = cats.get(item.category_id)
        item.category_name = cat.name_ko if cat else None
        item.category_color = cat.color_hex if cat else None
    return items

def attach_sop_coverage(db: Session, items: list) -> list:
    coverage = sop_coverage.get_standard_coverage(db, [item.id for item in items])
    for item in items:
        c = coverage.get(item.id, {"status": "없음", "count": 0})
        item.sop_status = c["status"]
        item.sop_count = c["count"]
    return items

def _track_changes(item: StandardItem, updates: dict, changed_by: int) -> list:
    histories = []
    for field, new_val in updates.items():
        old_val = getattr(item, field, None)
        if str(old_val) != str(new_val):
            histories.append(StandardHistory(
                standard_item_id=item.id,
                field_name=field,
                old_value=str(old_val) if old_val is not None else None,
                new_value=str(new_val) if new_val is not None else None,
                changed_by=changed_by,
            ))
    return histories
