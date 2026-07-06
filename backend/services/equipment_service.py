from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from fastapi import HTTPException
from datetime import date
from io import BytesIO

from models.equipment import Equipment, EquipmentCalibration, EquipmentStandardMapping, EquipmentInvestment
from schemas.equipment import EquipmentCreate, EquipmentUpdate, CalibrationCreate, InvestmentCreate
from services import pagination_helper


def _attach_expiry(item: Equipment) -> Equipment:
    """최신 교정 만료일 및 D-Day 계산 (calibrations가 로드된 상태에서 호출)"""
    valid = [c for c in item.calibrations if c.next_due_date]
    if valid:
        latest = max(valid, key=lambda c: c.next_due_date)
        item.latest_expiry = latest.next_due_date
        item.days_to_expiry = (latest.next_due_date - date.today()).days
    else:
        item.latest_expiry = None
        item.days_to_expiry = None
    return item


# ── 장비 CRUD ─────────────────────────────────────────
def list_equipment(
    db: Session, page: int, size: int,
    search: str | None, status: str | None, category: str | None
) -> dict:
    q = db.query(Equipment).options(joinedload(Equipment.calibrations))
    if search:
        q = q.filter(or_(
            Equipment.name.ilike(f"%{search}%"),
            Equipment.model.ilike(f"%{search}%"),
            Equipment.serial_number.ilike(f"%{search}%"),
        ))
    if status:
        q = q.filter(Equipment.status == status)
    if category:
        q = q.filter(Equipment.category == category)

    result = pagination_helper.paginate(q.order_by(Equipment.name), page, size)
    for item in result["items"]:
        _attach_expiry(item)
    return result


def get_equipment(db: Session, eq_id: int) -> Equipment:
    item = (
        db.query(Equipment)
        .options(
            joinedload(Equipment.calibrations),
            joinedload(Equipment.investments),
        )
        .filter(Equipment.id == eq_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="장비를 찾을 수 없습니다")
    _attach_expiry(item)
    return item


def create_equipment(db: Session, body: EquipmentCreate) -> Equipment:
    item = Equipment(**body.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    item.calibrations = []
    item.investments = []
    _attach_expiry(item)
    return item


def update_equipment(db: Session, eq_id: int, body: EquipmentUpdate) -> Equipment:
    item = get_equipment(db, eq_id)
    for k, v in body.model_dump().items():
        setattr(item, k, v)
    db.commit()
    db.refresh(item)
    _attach_expiry(item)
    return item


def delete_equipment(db: Session, eq_id: int):
    item = get_equipment(db, eq_id)
    db.delete(item)
    db.commit()


# ── 교정 이력 ─────────────────────────────────────────
def add_calibration(db: Session, eq_id: int, body: CalibrationCreate) -> EquipmentCalibration:
    get_equipment(db, eq_id)
    record = EquipmentCalibration(equipment_id=eq_id, **body.model_dump())
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def update_calibration(
    db: Session, eq_id: int, cal_id: int, body: CalibrationCreate
) -> EquipmentCalibration:
    record = (
        db.query(EquipmentCalibration)
        .filter(EquipmentCalibration.id == cal_id, EquipmentCalibration.equipment_id == eq_id)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="교정 이력을 찾을 수 없습니다")
    for k, v in body.model_dump().items():
        setattr(record, k, v)
    db.commit()
    db.refresh(record)
    return record


def delete_calibration(db: Session, eq_id: int, cal_id: int):
    record = (
        db.query(EquipmentCalibration)
        .filter(EquipmentCalibration.id == cal_id, EquipmentCalibration.equipment_id == eq_id)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="교정 이력을 찾을 수 없습니다")
    db.delete(record)
    db.commit()


# ── 규격 Capability 매핑 ─────────────────────────────────
def get_standard_item_ids(db: Session, eq_id: int) -> list[int]:
    get_equipment(db, eq_id)
    rows = db.query(EquipmentStandardMapping).filter(EquipmentStandardMapping.equipment_id == eq_id).all()
    return [r.standard_item_id for r in rows]


def set_standard_items(db: Session, eq_id: int, standard_item_ids: list[int]) -> dict:
    get_equipment(db, eq_id)
    db.query(EquipmentStandardMapping).filter(EquipmentStandardMapping.equipment_id == eq_id).delete()
    for sid in standard_item_ids:
        db.add(EquipmentStandardMapping(equipment_id=eq_id, standard_item_id=sid))
    db.commit()
    return {"ok": True}


# ── 투자 계획 ─────────────────────────────────────────
def list_investments(
    db: Session, year: int | None, equipment_id: int | None
) -> list[EquipmentInvestment]:
    q = db.query(EquipmentInvestment)
    if year:
        q = q.filter(EquipmentInvestment.year == year)
    if equipment_id:
        q = q.filter(EquipmentInvestment.equipment_id == equipment_id)
    records = q.order_by(EquipmentInvestment.year, EquipmentInvestment.invest_type).all()

    # equipment_name 주입
    eq_map: dict[int, str] = {}
    for r in records:
        if r.equipment_id and r.equipment_id not in eq_map:
            eq = db.query(Equipment).filter(Equipment.id == r.equipment_id).first()
            if eq:
                eq_map[r.equipment_id] = eq.name
    for r in records:
        r.equipment_name = eq_map.get(r.equipment_id) if r.equipment_id else None
    return records


def create_investment(db: Session, body: InvestmentCreate) -> EquipmentInvestment:
    record = EquipmentInvestment(**body.model_dump())
    db.add(record)
    db.commit()
    db.refresh(record)
    if record.equipment_id:
        eq = db.query(Equipment).filter(Equipment.id == record.equipment_id).first()
        record.equipment_name = eq.name if eq else None
    else:
        record.equipment_name = None
    return record


def update_investment(db: Session, inv_id: int, body: InvestmentCreate) -> EquipmentInvestment:
    record = db.query(EquipmentInvestment).filter(EquipmentInvestment.id == inv_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="투자 계획을 찾을 수 없습니다")
    for k, v in body.model_dump().items():
        setattr(record, k, v)
    db.commit()
    db.refresh(record)
    if record.equipment_id:
        eq = db.query(Equipment).filter(Equipment.id == record.equipment_id).first()
        record.equipment_name = eq.name if eq else None
    else:
        record.equipment_name = None
    return record


def delete_investment(db: Session, inv_id: int):
    record = db.query(EquipmentInvestment).filter(EquipmentInvestment.id == inv_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="투자 계획을 찾을 수 없습니다")
    db.delete(record)
    db.commit()


# ── 대시보드용: 교정 만료 임박 장비 ─────────────────────
def get_calibration_alerts(db: Session, days: int = 60) -> list[dict]:
    """만료 D-{days} 이내 장비 반환"""
    items = (
        db.query(Equipment)
        .options(joinedload(Equipment.calibrations))
        .filter(Equipment.status != "폐기")
        .all()
    )
    alerts = []
    today = date.today()
    for item in items:
        _attach_expiry(item)
        if item.days_to_expiry is not None and item.days_to_expiry <= days:
            alerts.append({
                "id": item.id,
                "name": item.name,
                "category": item.category,
                "status": item.status,
                "latest_expiry": item.latest_expiry,
                "days_to_expiry": item.days_to_expiry,
            })
    return sorted(alerts, key=lambda x: x["days_to_expiry"])


# ── 교정이력 관리 양식 ────────────────────────────────────
def generate_calibration_template() -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "교정이력"

    headers = ["장비명 *", "자산번호", "교정구분 *", "교정일자 *", "다음교정예정일", "결과 *", "교정기관", "성적서번호", "메모"]
    col_widths = [22, 14, 12, 14, 16, 12, 20, 18, 26]

    hdr_fill = PatternFill("solid", fgColor="2B2F82")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    for col, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 26

    row_fill = PatternFill("solid", fgColor="F5F7FA")
    samples = [
        ["항온항습챔버 #1", "AST-2019-014", "정기교정", "2026-03-15", "2027-03-14", "합격", "한국계량측정협회", "KC-2026-0312", ""],
        ["진동시험기 #2", "AST-2020-027", "정기교정", "2026-05-02", "2027-05-01", "합격", "KOLAS 인정기관", "KC-2026-0489", ""],
    ]
    for r, row_data in enumerate(samples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill

    ws2 = wb.create_sheet("작성 가이드")
    ws2["A1"] = "교정이력 관리 Excel 양식 작성 가이드"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 85
    for i, text in enumerate([
        "■ 장비명 *: 장비대장에 등록된 장비명과 동일하게 입력  ← 필수 입력",
        "■ 자산번호: 사내 자산관리 번호 (선택)",
        "■ 교정구분 *: 정기교정 / 특별교정 / 기능점검 중 선택  ← 필수 입력",
        "■ 교정일자 *: YYYY-MM-DD 형식  ← 필수 입력",
        "■ 다음교정예정일: YYYY-MM-DD 형식, 만료 알림(D-60) 기준일",
        "■ 결과 *: 합격 / 불합격 / 조건부합격 중 선택  ← 필수 입력",
        "■ 교정기관: 교정을 수행한 기관명",
        "■ 성적서번호: 교정 성적서 발급 번호",
        "■ 메모: 특이사항",
        "",
        "※ 1행(헤더)은 수정하지 마세요.",
        "※ 본 양식은 참고용 기록 양식입니다 — 시스템 일괄 등록(Import)은 미지원, 개별 입력은 장비 상세 화면에서 진행하세요.",
    ], 3):
        ws2.cell(row=i, column=1, value=text)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
